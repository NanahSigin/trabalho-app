import tkinter as tk
from tkinter import ttk, messagebox
import csv
import os

# Tenta importar openpyxl para salvar .xlsx com ajuste de colunas
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ============================
#   APLICATIVO IMOBILIÁRIA R.M
# ============================

class RMApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Imobiliária R.M")
        self.root.geometry("800x750")
        self.root.config(bg="#f0f2f5")
        self.root.resizable(False, False)

        self.usuarios = {}
        self.usuario_atual = None
        self.orcamento_dados = []

        # Estilos ttk
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TCombobox", fieldbackground="#ffffff", background="#ffffff", foreground="#333")
        self.style.configure("TSpinbox", fieldbackground="#ffffff", background="#ffffff", foreground="#333")

        self.frame = tk.Frame(root, bg="#f0f2f5")
        self.frame.pack(fill="both", expand=True)

        self.tela_cadastro()

    # =========================
    # FUNÇÕES DE BOTÃO COM HOVER
    # =========================
    def criar_botao(self, parent, texto, bg, fg, comando):
        btn = tk.Button(parent, text=texto, bg=bg, fg=fg, font=("Helvetica", 12, "bold"),
                        bd=0, relief="flat", activebackground=self.escurecer_cor(bg, 0.85),
                        activeforeground="white", command=comando, cursor="hand2")
        btn.bind("<Enter>", lambda e: btn.config(bg=self.escurecer_cor(bg, 0.85)))
        btn.bind("<Leave>", lambda e: btn.config(bg=bg))
        return btn

    def escurecer_cor(self, hex_color, fator=0.9):
        hex_color = hex_color.lstrip("#")
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        r = int(r * fator)
        g = int(g * fator)
        b = int(b * fator)
        return f"#{r:02x}{g:02x}{b:02x}"

    # =========================
    #   TELA DE CADASTRO
    # =========================
    def tela_cadastro(self):
        self.limpar_frame()

        card = tk.Frame(self.frame, bg="#ffffff")
        card.place(relx=0.5, rely=0.5, anchor="center", width=450, height=350)
        card.configure(highlightbackground="#cccccc", highlightthickness=1)
        card.pack_propagate(False)

        tk.Label(card, text="Cadastro de Usuário", bg="#ffffff", fg="#1a535c",
                 font=("Helvetica", 22, "bold")).pack(pady=(30, 20))

        tk.Label(card, text="Nome:", bg="#ffffff", font=("Helvetica", 13)).pack(pady=(5, 0), anchor="w", padx=30)
        self.nome_entry = tk.Entry(card, width=32, font=("Helvetica", 12), bd=2, relief="groove")
        self.nome_entry.pack(pady=(0, 10), padx=30, ipady=6)

        tk.Label(card, text="Senha:", bg="#ffffff", font=("Helvetica", 13)).pack(pady=(5, 0), anchor="w", padx=30)
        self.senha_entry = tk.Entry(card, show="*", width=32, font=("Helvetica", 12), bd=2, relief="groove")
        self.senha_entry.pack(pady=(0, 20), padx=30, ipady=6)

        btn = self.criar_botao(card, "Cadastrar", "#1a535c", "white", self.cadastrar_usuario)
        btn.pack(pady=10, ipadx=10, ipady=6)

    def cadastrar_usuario(self):
        nome = self.nome_entry.get().strip()
        senha = self.senha_entry.get().strip()
        if not nome or not senha:
            messagebox.showwarning("Atenção", "Preencha nome e senha!")
            return
        if nome in self.usuarios:
            messagebox.showwarning("Atenção", "Usuário já existe!")
            return
        self.usuarios[nome] = senha
        self.usuario_atual = nome
        messagebox.showinfo("Sucesso", f"Usuário '{nome}' cadastrado com sucesso!")
        self.tela_orcamento()

    # =========================
    #   TELA DE ORÇAMENTO
    # =========================
    def tela_orcamento(self):
        self.limpar_frame()

        # Cabeçalho gradiente
        header = tk.Canvas(self.frame, width=800, height=120, bg="#f0f2f5", highlightthickness=0)
        header.pack()
        for i in range(0, 800):
            cor = self.interpolar_cor("#1a535c", "#4ecdc4", i / 800)
            header.create_line(i, 0, i, 120, fill=cor)
        header.create_text(400, 60, text=f"Bem-vindo(a), {self.usuario_atual}", font=("Helvetica", 20, "bold"), fill="white")

        # Card principal
        card = tk.Frame(self.frame, bg="#ffffff")
        card.place(relx=0.5, rely=0.6, anchor="center", width=720, height=540)
        card.configure(highlightbackground="#cccccc", highlightthickness=1)
        card.pack_propagate(False)

        tk.Label(card, text="Gerador de Orçamento de Aluguel", bg="#ffffff",
                 font=("Helvetica", 16, "bold"), fg="#17252a").pack(pady=15)

        form_frame = tk.Frame(card, bg="#ffffff")
        form_frame.pack(pady=10, padx=20, fill="x")

        # Campos
        tk.Label(form_frame, text="Tipo de imóvel:", bg="#ffffff", font=("Helvetica", 12)).grid(row=0, column=0, sticky="w", pady=8)
        self.tipo_var = tk.StringVar(value="Apartamento")
        tipo_combo = ttk.Combobox(form_frame, textvariable=self.tipo_var, values=["Apartamento", "Casa", "Estúdio"], state="readonly", width=25)
        tipo_combo.grid(row=0, column=1, sticky="w", padx=10)

        tk.Label(form_frame, text="Quantidade de quartos:", bg="#ffffff", font=("Helvetica", 12)).grid(row=1, column=0, sticky="w", pady=8)
        self.quartos_var = tk.IntVar(value=1)
        quartos_spin = ttk.Spinbox(form_frame, from_=1, to=2, textvariable=self.quartos_var, width=7)
        quartos_spin.grid(row=1, column=1, sticky="w", padx=10)

        tk.Label(form_frame, text="Possui garagem/estacionamento:", bg="#ffffff", font=("Helvetica", 12)).grid(row=2, column=0, sticky="w", pady=8)
        self.garagem_var = tk.StringVar(value="N")
        garagem_combo = ttk.Combobox(form_frame, textvariable=self.garagem_var, values=["S","N"], state="readonly", width=25)
        garagem_combo.grid(row=2, column=1, sticky="w", padx=10)

        tk.Label(form_frame, text="Possui crianças:", bg="#ffffff", font=("Helvetica", 12)).grid(row=3, column=0, sticky="w", pady=8)
        self.criancas_var = tk.StringVar(value="S")
        criancas_combo = ttk.Combobox(form_frame, textvariable=self.criancas_var, values=["S","N"], state="readonly", width=25)
        criancas_combo.grid(row=3, column=1, sticky="w", padx=10)

        tk.Label(form_frame, text="Parcelas do contrato (1 a 5):", bg="#ffffff", font=("Helvetica", 12)).grid(row=4, column=0, sticky="w", pady=8)
        self.parcelas_var = tk.IntVar(value=1)
        parcelas_spin = ttk.Spinbox(form_frame, from_=1, to=5, textvariable=self.parcelas_var, width=7)
        parcelas_spin.grid(row=4, column=1, sticky="w", padx=10)

        tk.Label(form_frame, text="Vagas extras (somente Estúdio):", bg="#ffffff", font=("Helvetica", 12)).grid(row=5, column=0, sticky="w", pady=8)
        self.vagas_var = tk.IntVar(value=0)
        self.vagas_spinbox = ttk.Spinbox(form_frame, from_=0, to=2, textvariable=self.vagas_var, width=7, state="disabled")
        self.vagas_spinbox.grid(row=5, column=1, sticky="w", padx=10)

        # Atualização dinâmica dos campos
        def atualizar_campos(*args):
            tipo = self.tipo_var.get()
            # Vagas extras habilitadas somente para Estúdio
            if tipo != "Estúdio":
                self.vagas_spinbox.config(state="disabled")
                self.vagas_var.set(0)
            else:
                self.vagas_spinbox.config(state="normal")
                if self.vagas_var.get() > 2:
                    self.vagas_var.set(2)
            # Limitar quartos até 2 (spinbox já limita, mas reforçamos)
            if self.quartos_var.get() > 2:
                self.quartos_var.set(2)
            # Limitar parcelas até 5 (spinbox já limita, mas reforçamos)
            if self.parcelas_var.get() > 5:
                self.parcelas_var.set(5)

        self.tipo_var.trace("w", atualizar_campos)
        self.vagas_var.trace("w", atualizar_campos)
        self.quartos_var.trace("w", atualizar_campos)
        self.parcelas_var.trace("w", atualizar_campos)

        # Botões
        btn_frame = tk.Frame(card, bg="#ffffff")
        btn_frame.pack(pady=25)
        btn_gerar = self.criar_botao(btn_frame, "Gerar Orçamento", "#4ecdc4", "white", self.gerar_orcamento)
        btn_gerar.grid(row=0, column=0, padx=15, ipadx=15, ipady=6)
        btn_salvar = self.criar_botao(btn_frame, "Salvar Orçamento (.xlsx/.csv)", "#17252a", "white", self.salvar_csv)
        btn_salvar.grid(row=0, column=1, padx=15, ipadx=15, ipady=6)

        # Resultado
        self.result_label = tk.Label(card, text="", bg="#ffffff", font=("Helvetica", 12), justify="left")
        self.result_label.pack(pady=10, fill="x", padx=20)

    # =========================
    # Interpolação de cor para gradiente
    # =========================
    def interpolar_cor(self, cor1, cor2, fator):
        cor1 = cor1.lstrip("#")
        cor2 = cor2.lstrip("#")
        r = int(int(cor1[0:2],16)*(1-fator) + int(cor2[0:2],16)*fator)
        g = int(int(cor1[2:4],16)*(1-fator) + int(cor2[2:4],16)*fator)
        b = int(int(cor1[4:6],16)*(1-fator) + int(cor2[4:6],16)*fator)
        return f"#{r:02x}{g:02x}{b:02x}"

    # =========================
    # Formata número float para string no formato brasileiro: 1.234,56
    # =========================
    def formatar_brl(self, value):
        # value: número float
        s = f"{value:,.2f}"              # ex: '1,234.56'
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")  # '1.234,56'
        return f"R$ {s}"

    # =========================
    # FUNÇÕES ORÇAMENTO
    # =========================
    def gerar_orcamento(self):
        tipo = self.tipo_var.get()
        quartos = self.quartos_var.get()
        garagem = self.garagem_var.get()
        criancas = self.criancas_var.get()
        parcelas = self.parcelas_var.get()
        vagas_extras = self.vagas_var.get()

        valor_base = 700 if tipo == "Apartamento" else 900 if tipo == "Casa" else 1200
        if tipo in ["Apartamento", "Casa"] and quartos == 2:
            valor_base += 200 if tipo == "Apartamento" else 250
        if tipo == "Estúdio":
            if garagem == "S":
                valor_base += 250 + vagas_extras * 60
        if garagem == "S" and tipo != "Estúdio":
            valor_base += 300
        if tipo == "Apartamento" and criancas == "N":
            valor_base *= 0.95

        contrato = 2000
        parcelas = max(1, min(parcelas, 5))
        valor_parcela = contrato / parcelas
        total_mes = valor_base + valor_parcela

        texto = (f"🏡 Tipo: {tipo}\n"
                 f"💰 Aluguel mensal: R$ {valor_base:.2f}\n"
                 f"📜 Contrato: R$ {contrato:.2f} ({parcelas}x de R$ {valor_parcela:.2f})\n"
                 f"💵 Total do primeiro mês: R$ {total_mes:.2f}")
        self.result_label.config(text=texto)

        acumulado = 0
        self.orcamento_dados = []
        for mes in range(1, 13):
            parcela_mes = valor_parcela if mes <= parcelas else 0
            total_mes = valor_base + parcela_mes
            acumulado += total_mes
            self.orcamento_dados.append([mes, valor_base, parcela_mes, total_mes, acumulado])

    # =========================
    # SALVAR CSV / XLSX (com ajuste de colunas para evitar '#####')
    # =========================
    def salvar_csv(self):
        if not self.orcamento_dados:
            messagebox.showwarning("Atenção", "Gere o orçamento antes de salvar!")
            return

        downloads = os.path.join("C:\\Users\\anama\\Downloads")
        if not os.path.exists(downloads):
            os.makedirs(downloads)

        # Primeiro, tente salvar .xlsx (melhor experiência no Excel)
        if OPENPYXL_AVAILABLE:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Orçamento"

                headers = ["Mês", "Aluguel (R$)", "Parcela Contrato (R$)", "Total Mensal (R$)", "Acumulado (R$)"]
                ws.append(headers)

                # Inserir linhas formatadas em BRL (texto) — isso evita #### por causa de formatação inesperada.
                for linha in self.orcamento_dados:
                    mes, aluguel, parcela_contrato, total_mes, acumulado = linha
                    ws.append([
                        mes,
                        self.formatar_brl(aluguel),
                        self.formatar_brl(parcela_contrato),
                        self.formatar_brl(total_mes),
                        self.formatar_brl(acumulado)
                    ])

                # Ajustar largura das colunas com base no maior conteúdo de cada coluna
                for col_idx, col in enumerate(ws.columns, start=1):
                    max_length = 0
                    for cell in col:
                        try:
                            cell_value = str(cell.value) if cell.value is not None else ""
                        except Exception:
                            cell_value = ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    adjusted_width = (max_length + 2)
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = adjusted_width

                caminho_xlsx = os.path.join(downloads, "Orcamento_Aluguel.xlsx")
                wb.save(caminho_xlsx)
                messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{caminho_xlsx}")
                return
            except Exception as e:
                # se falhar, tenta salvar CSV como fallback
                print("Erro ao salvar .xlsx:", e)
                # continua para salvar CSV abaixo

        # Fallback: salvar CSV com valores formatados (R$ 1.234,56)
        try:
            caminho_csv = os.path.join(downloads, "Orcamento_Aluguel.csv")
            with open(caminho_csv, "w", newline="", encoding="utf-8-sig") as f:
                escritor = csv.writer(f, delimiter=";")
                escritor.writerow(["Mês", "Aluguel (R$)", "Parcela Contrato (R$)", "Total Mensal (R$)", "Acumulado (R$)"])
                for linha in self.orcamento_dados:
                    mes, aluguel, parcela_contrato, total_mes, acumulado = linha
                    escritor.writerow([
                        mes,
                        self.formatar_brl(aluguel),
                        self.formatar_brl(parcela_contrato),
                        self.formatar_brl(total_mes),
                        self.formatar_brl(acumulado)
                    ])
            messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{caminho_csv}\n\nDica: abra o arquivo no Excel e ajuste as colunas (duplo-clique entre cabeçalhos) se necessário.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar CSV:\n{e}")

    # =========================
    # LIMPAR FRAME
    # =========================
    def limpar_frame(self):
        for widget in self.frame.winfo_children():
            widget.destroy()


# =========================
# EXECUÇÃO
# =========================
if __name__ == "__main__":
    root = tk.Tk()
    app = RMApp(root)
    root.mainloop()
